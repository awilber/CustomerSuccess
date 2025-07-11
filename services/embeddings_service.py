import os
import json
import hashlib
from datetime import datetime
# import numpy as np  # Optional, using math instead
from typing import List, Dict, Optional
import re
from collections import Counter
# Optional imports - gracefully handle if not available
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None
try:
    import docx
except ImportError:
    docx = None
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None
from models import db, FileReference, EmailThread

class EmbeddingsService:
    """Service for generating embeddings and analyzing file content"""
    
    def __init__(self):
        # Simple TF-IDF based embeddings for now
        # In production, you'd use sentence-transformers or OpenAI embeddings
        self.vocabulary = {}
        self.idf_scores = {}
        self.stopwords = set([
            'the', 'is', 'at', 'which', 'on', 'a', 'an', 'and', 'or', 'but',
            'in', 'with', 'to', 'for', 'of', 'as', 'by', 'that', 'this',
            'it', 'from', 'be', 'are', 'been', 'was', 'were', 'being'
        ])
    
    def process_file(self, file_ref_id):
        """Process a file to extract content and generate embeddings"""
        file_ref = FileReference.query.get(file_ref_id)
        if not file_ref or not os.path.exists(file_ref.file_path):
            return False
        
        try:
            file_ref.processing_status = 'processing'
            db.session.commit()
            
            # Extract text content
            content = self._extract_content(file_ref.file_path, file_ref.file_type)
            
            # Generate summary
            summary = self._generate_summary(content)
            
            # Extract keywords
            keywords = self._extract_keywords(content)
            
            # Detect topic
            topic = self._detect_topic(content, keywords)
            
            # Generate embedding
            embedding = self._generate_embedding(content)
            
            # Update file reference
            file_ref.summary = summary
            file_ref.keywords = json.dumps(keywords)
            file_ref.topic = topic
            file_ref.embedding = json.dumps(embedding)
            file_ref.processing_status = 'completed'
            file_ref.processed_at = datetime.utcnow()
            
            db.session.commit()
            return True
            
        except Exception as e:
            file_ref.processing_status = 'error'
            file_ref.error_message = str(e)
            db.session.commit()
            return False
    
    def _extract_content(self, filepath, file_type):
        """Extract text content from various file types"""
        content = ""
        
        try:
            if file_type in ['txt', 'md', 'log']:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    
            elif file_type == 'pdf':
                content = self._extract_pdf_content(filepath)
                
            elif file_type in ['doc', 'docx']:
                content = self._extract_docx_content(filepath)
                
            elif file_type in ['xls', 'xlsx']:
                content = self._extract_excel_content(filepath)
                
            elif file_type in ['html', 'htm']:
                content = self._extract_html_content(filepath)
                
            elif file_type in ['py', 'js', 'java', 'cpp', 'c', 'cs', 'rb', 'go']:
                # For code files, extract comments and docstrings
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = self._extract_code_documentation(f.read(), file_type)
                    
            elif file_type in ['json', 'xml', 'yaml', 'yml']:
                # For structured data, extract keys and values
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()[:10000]  # Limit size
                    
        except Exception as e:
            print(f"Error extracting content from {filepath}: {str(e)}")
        
        return content[:50000]  # Limit content size
    
    def _extract_pdf_content(self, filepath):
        """Extract text from PDF files"""
        if not PyPDF2:
            return "[PDF content - install PyPDF2 to extract]"
        content = ""
        try:
            with open(filepath, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages[:50]:  # Limit to first 50 pages
                    content += page.extract_text() + "\n"
        except:
            pass
        return content
    
    def _extract_docx_content(self, filepath):
        """Extract text from Word documents"""
        if not docx:
            return "[Word document - install python-docx to extract]"
        content = ""
        try:
            doc = docx.Document(filepath)
            for paragraph in doc.paragraphs[:500]:  # Limit paragraphs
                content += paragraph.text + "\n"
        except:
            pass
        return content
    
    def _extract_excel_content(self, filepath):
        """Extract text from Excel files"""
        if not openpyxl:
            return "[Excel content - install openpyxl to extract]"
        content = ""
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            for sheet_name in workbook.sheetnames[:5]:  # First 5 sheets
                sheet = workbook[sheet_name]
                content += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(max_row=100, values_only=True):
                    content += " ".join(str(cell) for cell in row if cell) + "\n"
        except:
            pass
        return content
    
    def _extract_html_content(self, filepath):
        """Extract text from HTML files"""
        if not BeautifulSoup:
            # Fallback to regex-based extraction
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    # Remove HTML tags
                    content = re.sub(r'<[^>]+>', ' ', content)
                    return content
            except:
                return ""
        content = ""
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                content = soup.get_text()
        except:
            pass
        return content
    
    def _extract_code_documentation(self, code, file_type):
        """Extract comments and documentation from code"""
        content = ""
        
        # Extract comments based on file type
        if file_type in ['py']:
            # Python docstrings and comments
            docstrings = re.findall(r'"""(.*?)"""', code, re.DOTALL)
            comments = re.findall(r'#\s*(.*?)$', code, re.MULTILINE)
            content = "\n".join(docstrings + comments)
            
        elif file_type in ['js', 'java', 'cpp', 'c', 'cs', 'go']:
            # C-style comments
            block_comments = re.findall(r'/\*(.*?)\*/', code, re.DOTALL)
            line_comments = re.findall(r'//\s*(.*?)$', code, re.MULTILINE)
            content = "\n".join(block_comments + line_comments)
        
        return content
    
    def _generate_summary(self, content):
        """Generate a summary of the content"""
        if not content:
            return "Empty file"
        
        # Simple extractive summary - take first few sentences
        sentences = re.split(r'[.!?]+', content)
        sentences = [s.strip() for s in sentences if len(s.strip()) > 20]
        
        if not sentences:
            return content[:200] + "..." if len(content) > 200 else content
        
        return ". ".join(sentences[:3]) + "."
    
    def _extract_keywords(self, content):
        """Extract keywords from content"""
        if not content:
            return []
        
        # Tokenize and clean
        words = re.findall(r'\b[a-zA-Z]{3,}\b', content.lower())
        words = [w for w in words if w not in self.stopwords]
        
        # Get word frequencies
        word_freq = Counter(words)
        
        # Return top keywords
        keywords = [word for word, count in word_freq.most_common(20)]
        return keywords
    
    def _detect_topic(self, content, keywords):
        """Detect the main topic of the content"""
        topic_keywords = {
            'contract': ['agreement', 'contract', 'terms', 'conditions', 'clause', 'party', 'obligations'],
            'technical': ['api', 'code', 'function', 'error', 'bug', 'system', 'database', 'server'],
            'financial': ['payment', 'invoice', 'cost', 'budget', 'expense', 'revenue', 'profit'],
            'project': ['timeline', 'milestone', 'deliverable', 'schedule', 'deadline', 'phase'],
            'design': ['design', 'layout', 'interface', 'mockup', 'wireframe', 'ui', 'ux'],
            'meeting': ['meeting', 'minutes', 'agenda', 'discussion', 'action', 'notes'],
            'report': ['report', 'analysis', 'summary', 'findings', 'results', 'conclusion']
        }
        
        # Score each topic based on keyword matches
        topic_scores = {}
        content_lower = content.lower()
        
        for topic, topic_words in topic_keywords.items():
            score = sum(1 for word in topic_words if word in content_lower)
            score += sum(0.5 for word in topic_words if word in keywords)
            topic_scores[topic] = score
        
        # Return highest scoring topic
        best_topic = max(topic_scores.items(), key=lambda x: x[1])
        return best_topic[0] if best_topic[1] > 0 else 'general'
    
    def _generate_embedding(self, content):
        """Generate a simple embedding vector"""
        if not content:
            return [0.0] * 100
        
        # Simple approach: TF-IDF based embedding
        words = re.findall(r'\b[a-zA-Z]{3,}\b', content.lower())
        words = [w for w in words if w not in self.stopwords]
        
        # Create a simple hash-based embedding
        embedding = [0.0] * 100
        for word in set(words):
            # Hash word to position
            hash_val = int(hashlib.md5(word.encode()).hexdigest(), 16)
            position = hash_val % 100
            # TF score
            tf = words.count(word) / len(words) if words else 0
            embedding[position] += tf
        
        # Normalize without numpy
        norm = sum(x*x for x in embedding) ** 0.5
        if norm > 0:
            embedding = [x / norm for x in embedding]
        
        return embedding
    
    def calculate_similarity(self, embedding1, embedding2):
        """Calculate cosine similarity between two embeddings"""
        if isinstance(embedding1, str):
            embedding1 = json.loads(embedding1)
        if isinstance(embedding2, str):
            embedding2 = json.loads(embedding2)
        
        dot_product = sum(a * b for a, b in zip(embedding1, embedding2))
        return dot_product  # Already normalized

# Global instance
embeddings_service = EmbeddingsService()